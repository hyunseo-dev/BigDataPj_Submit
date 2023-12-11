from bs4 import BeautifulSoup as bs
from pathlib import Path
from typing import Optional, Union, Dict, List
from openpyxl import Workbook
import time
import os
import re
import requests as rq
import json
import mysql.connector
import pandas as pd


def get_headers(
        key: str,
        default_value: Optional[str] = None
) -> Dict[str, Dict[str, str]]:
    """ Get Headers """
    JSON_FILE: str = 'json/headers.json'

    with open(JSON_FILE, 'r', encoding='UTF-8') as file:
        headers: Dict[str, Dict[str, str]] = json.loads(file.read())

    try:
        return headers[key]
    except:
        if default_value:
            return default_value
        raise EnvironmentError(f'Set the {key}')


class Coupang:
    @staticmethod
    def get_product_code(url: str) -> str:
        """ 입력받은 URL 주소의 PRODUCT CODE 추출하는 메소드 """
        prod_code: str = url.split('products/')[-1].split('?')[0]
        return prod_code

    def __init__(self) -> None:
        self.__headers: Dict[str, str] = get_headers(key='headers')

    def main(self, url: str, page_count: int) -> List[List[Dict[str, Union[str, int]]]]:
        # URL의 Product Code 추출
        prod_code: str = self.get_product_code(url=url)

        # URL 주소 재가공
        URLS: List[str] = [
            f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={page}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true'
            for page in range(1, page_count + 1)]

        # __headers에 referer 키 추가
        self.__headers['referer'] = url

        with rq.Session() as session:
            return [self.fetch(url=url, session=session) for url in URLS]

    def fetch(self, url: str, session) -> List[Dict[str, Union[str, int]]]:
        save_data: List[Dict[str, Union[str, int]]] = list()

        with session.get(url=url, headers=self.__headers) as response:
            html = response.text
            soup = bs(html, 'html.parser')

            # Article Boxes
            article_length = len(soup.select('article.sdp-review__article__list'))

            for idx in range(article_length):
                dict_data: Dict[str, Union[str, int]] = dict()
                articles = soup.select('article.sdp-review__article__list')

                # 구매자 이름
                user_name = articles[idx].select_one('span.sdp-review__article__list__info__user__name')
                if user_name == None or user_name.text == '':
                    user_name = '-'
                else:
                    user_name = user_name.text.strip()

                # 평점
                rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
                if rating == None:
                    rating = 0
                else:
                    rating = int(rating.attrs['data-rating'])

                # 구매자 상품명
                prod_name = articles[idx].select_one('div.sdp-review__article__list__info__product-info__name')
                if prod_name == None or prod_name.text == '':
                    prod_name = '-'
                else:
                    prod_name = prod_name.text.strip()

                # 헤드라인(타이틀)
                headline = articles[idx].select_one('div.sdp-review__article__list__headline')
                if headline == None or headline.text == '':
                    headline = '등록된 헤드라인이 없습니다'
                else:
                    headline = headline.text.strip()

                # 리뷰 내용
                review_content = articles[idx].select_one('div.sdp-review__article__list__review > div')
                if review_content == None:
                    review_content = '등록된 리뷰내용이 없습니다'
                else:
                    review_content = re.sub('[\n\t]', '', review_content.text.strip())

                # 만족도
                answer_pairs = articles[idx].select('.sdp-review__article__list__survey__row')
                answer_list = []

                for answer_pair in answer_pairs:
                    question = answer_pair.select_one('.sdp-review__article__list__survey__row__question')
                    answer = answer_pair.select_one('.sdp-review__article__list__survey__row__answer')

                    question_text = '질문 없음' if question is None or question.text == '' else question.text.strip()
                    answer_text = '평가 없음' if answer is None or answer.text == '' else answer.text.strip()

                    answer_list.append({'question': question_text, 'answer': answer_text})

                dict_data['prod_name'] = prod_name
                dict_data['user_name'] = user_name
                dict_data['rating'] = rating
                dict_data['headline'] = headline
                dict_data['review_content'] = review_content
                dict_data['answer'] = answer_list  # 여러 쌍의 질문과 답변을 리스트로 저장

                save_data.append(dict_data)

                print(dict_data, '\n')

            time.sleep(1)

            return save_data

    @staticmethod
    def clear_console() -> None:
        command: str = 'clear'
        if os.name in ('nt', 'dos'):
            command = 'cls'
        os.system(command=command)

    def input_page_count(self) -> int:
        self.clear_console()

        while True:
            page_count: str = input('페이지 수를 입력하세요\n\n:')
            if not page_count:
                print('페이지 수가 입력되지 않았습니다\n')
                continue

            return int(page_count)


class OpenPyXL:
    @staticmethod
    def save_file(results: List[List[Dict[str, Union[str, int]]]]) -> None:
        wb = Workbook()
        ws = wb.active

        # 엑셀 헤더에 필요한 열 추가
        header = ['상품명', '구매자 이름', '구매자 평점', '리뷰 제목', '리뷰 내용']
        if results:
            for answer_pair in results[0][0]['answer']:
                header.append(f'{answer_pair["question"]}')

        ws.append(header)

        row = 2

        for x in results:
            for result in x:
                ws[f'A{row}'] = result['prod_name']
                ws[f'B{row}'] = result['user_name']
                ws[f'C{row}'] = result['rating']
                ws[f'D{row}'] = result['headline']
                ws[f'E{row}'] = result['review_content']

                # 각 쌍의 만족도(question 및 answer)를 엑셀에 저장
                for idx, answer_pair in enumerate(result['answer'], start=1):
                    ws[f'{chr(69+idx)}{row}'] = answer_pair['answer']

                row += 1

        savePath: str = os.path.abspath('쿠팡상품리뷰')
        fileName: str = '상품1' + '.xlsx'

        if not os.path.exists(savePath):
            os.mkdir(savePath)

        wb.save(os.path.join(savePath, fileName))
        wb.close()

        print(f'파일 저장완료!\n\n{os.path.join(savePath, fileName)}')

def create_tables():
    try:
        # MySQL 연결 설정
        connection = mysql.connector.connect(
            host='127.0.0.1',
            user='root',
            password='password',
            database='reviews'
        )

        # 커서 생성
        cursor = connection.cursor()

        # '상품명' 테이블 생성
        create_product_table_query = '''
            CREATE TABLE IF NOT EXISTS `reviews`.`Product_Name` (
                `Product_name` VARCHAR(255) NOT NULL,
                `Color` VARCHAR(45) NULL,
                `Storage` VARCHAR(45) NULL,
                `Network_Type` VARCHAR(45) NULL
            )
        '''
        cursor.execute(create_product_table_query)

        # '유저 이름' 테이블 생성
        create_user_table_query = '''
            CREATE TABLE IF NOT EXISTS `reviews`.`UserName` (
                `UserName` VARCHAR(255) NOT NULL,
                PRIMARY KEY (`UserName`)
            )
        '''
        cursor.execute(create_user_table_query)

        # '별점' 테이블 생성
        create_rating_table_query = '''
            CREATE TABLE IF NOT EXISTS `reviews`.`Ratings` (
                `Ratings` INT NOT NULL
            )
        '''
        cursor.execute(create_rating_table_query)

        # '리뷰 제목' 테이블 생성
        create_review_title_table_query = '''
            CREATE TABLE IF NOT EXISTS `reviews`.`Review_Title` (
                `Title` VARCHAR(255) NOT NULL
            )
        '''
        cursor.execute(create_review_title_table_query)

        # '리뷰 내용' 테이블 생성
        create_review_content_table_query = '''
            CREATE TABLE IF NOT EXISTS `reviews`.`Review` (
                `Review` LONGTEXT NOT NULL
            )
        '''
        cursor.execute(create_review_content_table_query)

        # 변경사항 저장
        connection.commit()

        print('테이블이 성공적으로 생성되었습니다.')

    except mysql.connector.Error as err:
        print(f'오류: {err}')

    finally:
        # 연결 종료
        cursor.close()
        connection.close()

def insert_data_to_tables(excel_file_path: str):
    try:
        # MySQL 연결 설정
        connection = mysql.connector.connect(
            host='127.0.0.1',
            user='root',
            password='password',
            database='reviews'
        )

        # 커서 생성
        cursor = connection.cursor()

        df = pd.read_excel(excel_file_path, header=None, skiprows=1)  # 헤더를 생략하고 2번째 행부터 읽어옴

        # 각 행에 대해 데이터베이스에 삽입
        for index, row in df.iterrows():
            # 상품명 테이블에 데이터 삽입
            prod_name_query = '''
                INSERT INTO `Product_Name` (`Product_name`, `Color`, `Storage`, `Network_Type`)
                VALUES (%s, %s, %s, %s)
            '''
            # row[0]을 콤마로 구분하여 각 값을 추출
            product_info = row[0].split(',')

            # 각 값을 변수에 할당
            product_name = product_info[0].strip()  # strip() 함수는 문자열 양 끝의 공백을 제거합니다
            color = product_info[1].strip() if len(product_info) > 1 else None
            storage = product_info[2].strip() if len(product_info) > 2 else None
            network_type = product_info[3].strip() if len(product_info) > 3 else None

            # 값을 튜플로 구성
            prod_name_values = (
                product_name,
                color,
                storage,
                network_type
            )

            # 데이터베이스에 삽입
            cursor.execute(prod_name_query, prod_name_values)

            # 유저이름 테이블에 데이터 삽입
            user_name_query = '''
                INSERT INTO `UserName` (`UserName`)
                VALUES (%s)
            '''
            user_name_values = (
                row[1],
            )
            cursor.execute(user_name_query, user_name_values)

            # 별점 테이블에 데이터 삽입
            rating_query = '''
                INSERT INTO `Ratings` (`Ratings`)
                VALUES (%s)
            '''
            rating_values = (
                row[2],
            )
            cursor.execute(rating_query, rating_values)

            # 리뷰제목 테이블에 데이터 삽입
            review_title_query = '''
                INSERT INTO `Review_Title` (`Title`)
                VALUES (%s)
            '''
            review_title_values = (
                row[3],
            )
            cursor.execute(review_title_query, review_title_values)

            # 리뷰내용 테이블에 데이터 삽입
            review_content_query = '''
                INSERT INTO `Review` (`Review`)
                VALUES (%s)
            '''
            review_content_values = (
                row[4],
            )
            cursor.execute(review_content_query, review_content_values)

        # 변경사항 저장
        connection.commit()

        print('데이터 삽입이 완료되었습니다.')

    except mysql.connector.Error as err:
        print(f'오류: {err}')

    finally:
        # 연결 종료
        cursor.close()
        connection.close()


if __name__ == "__main__":
    coupang = Coupang()

    # 터미널에서 URL 입력 받기
    url = input("리뷰를 수집할 쿠팡 상품 URL을 입력하세요: ")

    # 페이지 수 입력 받기
    page_count = coupang.input_page_count()

    # 크롤링 실행
    results = coupang.main(url=url, page_count=page_count)

    # 결과를 엑셀 파일로 저장
    OpenPyXL.save_file(results=results)

    # 테이블 생성 함수 호출
    create_tables()

    # 테이블에 데이터 삽입 함수 호출 (엑셀 파일 경로는 적절하게 수정)
    insert_data_to_tables('C:/Users/hsbob/PycharmProjects/MySQL_Scratch/쿠팡상품리뷰/상품1.xlsx')
